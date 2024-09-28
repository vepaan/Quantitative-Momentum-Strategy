import pandas as pd
from openpyxl.styles import PatternFill

rsi_bollinger_file = 'portfolios/RSI_Bollinger.csv'
macd_file = 'portfolios/MACD.csv'
output_file = 'portfolios/portfolio.xlsx'
portfolio_data_file = 'data/processed_data/portfolio_data.csv'  

rsi_bollinger_df = pd.read_csv(rsi_bollinger_file)
macd_df = pd.read_csv(macd_file)
data = pd.read_csv(portfolio_data_file)

if 'Ticker' in rsi_bollinger_df.columns and 'Ticker' in macd_df.columns:
    rsi_bollinger_selected = rsi_bollinger_df[['Ticker', 'Date', 'Latest Close', 'Upper Band', 'Lower Band', 'RSI', 'RSI signal', 'Bollinger signal']]
    macd_selected = macd_df[['Ticker', 'Date', 'MACD', 'Signal', 'MACD Signal']]
    combined_df = pd.merge(rsi_bollinger_selected, macd_selected, on=['Ticker', 'Date'], how='outer', suffixes=('x', 'y'))
    combined_df = combined_df.loc[:, ~combined_df.columns.duplicated()]
else:
    print("One of the DataFrames does not contain the 'Ticker' column. Skipping merge.")
    combined_df = pd.concat([rsi_bollinger_df[['Ticker', 'Date']], macd_df[['Ticker', 'Date']]], axis=0, ignore_index=True)

momentum_data = []
for ticker in data['Ticker'].unique():
    close_data = data[data['Ticker'] == ticker]
    momentum = (close_data['Close'].iloc[-30]) / (close_data['Close'].iloc[0])
    momentum_data.append({'Ticker': ticker, 'Momentum': momentum})

momentum_df = pd.DataFrame(momentum_data)

combined_df = pd.merge(combined_df, momentum_df, on='Ticker', how='outer')

combined_df['Momentum Signal'] = combined_df['Momentum'].apply(
    lambda x: 'Buy' if x > 1 else 'Sell' if x < 1 else 'Hold'
)

# im reordering the buy signlas at last so its easier to compare
signal_columns = ['RSI signal', 'Bollinger signal', 'MACD Signal', 'Momentum Signal']
non_signal_columns = [col for col in combined_df.columns if col not in signal_columns]
ordered_columns = non_signal_columns + signal_columns
combined_df = combined_df[ordered_columns]

combined_df.to_csv(output_file.replace('.xlsx', '.csv'), index=False)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    combined_df.to_excel(writer, index=False, sheet_name='Portfolio')

    workbook = writer.book
    worksheet = writer.sheets['Portfolio']

    buy_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for Buy
    sell_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red for Sell
    hold_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for Hold

    for row in range(2, len(combined_df) + 2): 
        for col in signal_columns: 
            if col in combined_df.columns:
                cell_value = combined_df[col].iloc[row - 2]  
                cell = worksheet.cell(row=row, column=combined_df.columns.get_loc(col) + 1)  # Get the cell location
                cell.value = cell_value 

                if cell_value == 'Buy':
                    cell.fill = buy_fill
                elif cell_value == 'Sell':
                    cell.fill = sell_fill
                elif cell_value == 'Hold':
                    cell.fill = hold_fill

print(f"\nData successfully written to {output_file}.")
